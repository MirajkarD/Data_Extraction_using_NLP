import os
import nltk
from nltk.tokenize import word_tokenize, sent_tokenize
from nltk.corpus import stopwords
import string
import re
import openpyxl
from openpyxl.styles import Font

def load_words_from_file(file_path):
    """Load words from a file and return a set of unique words."""
    with open(file_path, 'r', encoding='latin-1') as file:
        words = file.read().split()
    return set(words)

def calculate_scores(text, positive_words, negative_words):
    """Calculate positive and negative scores based on dictionaries."""
    tokens = text.lower().split()
    tokens = [token for token in tokens if token]  # Remove empty tokens
    positive_score = sum(1 for token in tokens if token in positive_words)
    negative_score = -sum(1 for token in tokens if token in negative_words)
    # Multiply negative score by -1 to make it positive
    negative_score *= -1
    return positive_score, negative_score

# Function to calculate polarity score
def calculate_polarity_score(positive_score, negative_score):
    epsilon = 0.000001
    return (positive_score - negative_score) / (positive_score + negative_score + epsilon)

# Function to calculate subjectivity score
def calculate_subjectivity_score(positive_score, negative_score, total_words):
    epsilon = 0.000001
    return (positive_score + negative_score) / (total_words + epsilon)

# Function to calculate average sentence length
def avg_sentence_length(tokens):
    sentences = sent_tokenize(' '.join(tokens))
    total_words = len(tokens)
    return total_words / len(sentences)

# Function to calculate percentage of complex words
def percentage_complex_words(tokens):
    complex_word_count = sum(1 for word in tokens if len(word) > 2) # Assuming > 2 syllables is complex
    return (complex_word_count / len(tokens)) * 100

# Function to calculate FOG index
def fog_index(tokens):
    return 0.4 * (avg_sentence_length(tokens) + percentage_complex_words(tokens))

# Function to calculate average number of words per sentence
def avg_words_per_sentence(tokens):
    sentences = sent_tokenize(' '.join(tokens))
    total_words = len(tokens)
    return total_words / len(sentences)

# Function to count complex words
def count_complex_words(tokens):
    return sum(1 for word in tokens if len(word) > 2) # Assuming > 2 syllables is complex

# Function to count total words
def count_total_words(tokens):
    return len(tokens)

# Function to count syllables in a word
def count_syllables(word):
    vowels = 'aeiouy'
    count = 0
    prev_char_was_vowel = False
    for char in word:
        if char.lower() in vowels:
            if not prev_char_was_vowel:
                count += 1
                prev_char_was_vowel = True
        else:
            prev_char_was_vowel = False
    # Exception cases
    if word.endswith('es'):
        count -= 1
    if word.endswith('ed'):
        count -= 1
    if count == 0:
        count = 1
    return count

# Function to count personal pronouns
def count_personal_pronouns(text):
    pronouns = re.findall(r'\b(I|we|my|ours|us)\b', text, flags=re.IGNORECASE)
    return len(pronouns)

# Function to calculate average word length
def avg_word_length(tokens):
    total_chars = sum(len(word) for word in tokens)
    return total_chars / len(tokens)

# Extract url_Id from filename
def extract_url_id(filename):
    # Assuming url_Id is between underscores in the filename
    parts = filename.split('_')
    for part in parts:
        if part.startswith("url"):
            return part
    return None

# Load positive and negative words from files
positive_words_file = '../MasterDictionary/positive-words.txt'
negative_words_file = '../MasterDictionary/negative-words.txt'
positive_words = load_words_from_file(positive_words_file)
negative_words = load_words_from_file(negative_words_file)

# Path to the folder containing text files
folder_path = "../Data Extraction/output"

# Load URL_Id and URL from Excel file
excel_file_path = "Output Data Structure.xlsx"
wb = openpyxl.load_workbook(excel_file_path)
ws = wb.active

# Create a new workbook for output
output_wb = openpyxl.Workbook()
output_ws = output_wb.active

# Write headers
headers = [
    'URL_Id', 'URL', 'Positive Score', 'Negative Score', 'Polarity Score', 'Subjectivity Score',
    'Avg Sentence Length', 'Percentage of Complex Words', 'FOG Index',
    'Avg Number of Words per Sentence', 'Complex Word Count', 'Total Word Count', 'Syllables per Word',
    'Personal Pronouns', 'Avg Word Length'
]
for col, header in enumerate(headers, start=1):
    cell = output_ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)  # Make header bold

# Process each file in the folder
for row, filename in enumerate(os.listdir(folder_path), start=2):
    if filename.endswith(".txt"):
        file_path = os.path.join(folder_path, filename)
        with open(file_path, 'r', encoding='utf-8') as file:
            text = file.read()

        # Tokenize the text
        tokens = word_tokenize(text.lower())

        # Calculating variables
        positive_score, negative_score = calculate_scores(text, positive_words, negative_words)
        polarity_score = calculate_polarity_score(positive_score, negative_score)
        subjectivity_score = calculate_subjectivity_score(positive_score, negative_score, len(tokens))
        avg_sentence_len = avg_sentence_length(tokens)
        percentage_complex = percentage_complex_words(tokens)
        fog_index_val = fog_index(tokens)
        avg_words_per_sentence_val = avg_words_per_sentence(tokens)
        complex_word_count = count_complex_words(tokens)
        total_word_count = count_total_words(tokens)
        syllables_per_word = sum(count_syllables(word) for word in tokens) / len(tokens)
        personal_pronouns_count = count_personal_pronouns(text)
        avg_word_length_val = avg_word_length(tokens)

        # Extract URL_Id from filename
        URL_Id = ws.cell(row=row, column=1).value

        # Extract URL from Excel
        URL = ws.cell(row=row, column=2).value

        # Write results to output Excel
        output_ws.append([
            URL_Id, URL, positive_score, negative_score, polarity_score, subjectivity_score,
            avg_sentence_len, percentage_complex, fog_index_val,
            avg_words_per_sentence_val, complex_word_count, total_word_count, syllables_per_word,
            personal_pronouns_count, avg_word_length_val
        ])

# Save the output workbook
output_wb.save("output_features.xlsx")
print("Excel file with extracted features created successfully.")
